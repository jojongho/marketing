import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 엑셀 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "키워드_조합기"

# 1. 헤더 설정
headers = ["그룹1 (지역)", "그룹2 (업종/유형)", "그룹3 (세부/관심사)", "", "1차 조합 (1+2)", "2차 조합 (1+2+3)"]
ws.append(headers)

# 2. 예시 데이터 입력 (사용자 이해를 돕기 위함)
ws["A2"] = "아산"
ws["A3"] = "배방"
ws["B2"] = "아파트"
ws["B3"] = "분양"
ws["C2"] = "정보"
ws["C3"] = "시세"

# 3. 수식 입력 (2행부터 1000행까지)
# 엑셀 수식은 문자열로 입력
for row in range(2, 1002):
    # E열: 1차 조합 (A + B)
    # =INDEX(A:A, INT((ROW()-2)/COUNTA(B:B))+2) & " " & INDEX(B:B, MOD(ROW()-2, COUNTA(B:B))+2)
    # 주의: openpyxl에서 수식 입력 시 영어 함수명 사용
    formula_1 = f'=IFERROR(INDEX(A:A, INT((ROW()-2)/COUNTA(B:B))+2) & " " & INDEX(B:B, MOD(ROW()-2, COUNTA(B:B))+2), "")'
    ws[f"E{row}"] = formula_1

    # F열: 2차 조합 (E + C) -> E열(1차조합)을 메인으로 C열과 조합
    # =INDEX(E:E, INT((ROW()-2)/COUNTA(C:C))+2) & " " & INDEX(C:C, MOD(ROW()-2, COUNTA(C:C))+2)
    formula_2 = f'=IFERROR(INDEX(E:E, INT((ROW()-2)/COUNTA(C:C))+2) & " " & INDEX(C:C, MOD(ROW()-2, COUNTA(C:C))+2), "")'
    ws[f"F{row}"] = formula_2

from openpyxl.styles import Font

# 4. 스타일링 (헤더 강조)
for cell in ws[1]:
    cell.font = Font(bold=True)

# 5. 안내 문구 추가
ws["H2"] = "사용법:"
ws["H3"] = "1. A, B, C열에 원하는 키워드를 입력하세요."
ws["H4"] = "2. E열과 F열에 자동으로 조합된 키워드가 나타납니다."
ws["H5"] = "3. 결과가 안 보이면 수식을 더 아래로 드래그하세요."

# 파일 저장
output_path = "Keyword_Combiner_Tool.xlsx"
wb.save(output_path)
print(f"Created {output_path}")
