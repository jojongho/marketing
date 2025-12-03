"""
개선된 키워드 조합기 엑셀 생성 도구
- 실용적인 DB 시트 구조
- 자동 조합 수식
- 네이버 업로드용 CSV 자동 생성
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 워크북 생성
wb = Workbook()

# ============================================
# 1. 키워드 DB 시트들
# ============================================

# 1-1. 부동산 키워드 DB
ws_re = wb.create_sheet("DB_부동산", 0)
ws_re.append(["지역(대)", "지역(중)", "지역(소)", "건물유형", "거래유형", "타겟", "특성"])
ws_re.append(["충남", "아산시", "배방읍", "아파트", "분양", "실거주", "역세권"])
ws_re.append(["충남", "아산시", "탕정면", "오피스텔", "전세", "신혼부부", "초품아"])
ws_re.append(["충남", "천안시", "불당동", "빌라", "월세", "직장인", "학군"])
ws_re.append(["", "", "", "주상복합", "매매", "투자자", "신축"])
ws_re.append(["", "", "", "상가", "임대", "", ""])

# 1-2. 생활 키워드 DB
ws_life = wb.create_sheet("DB_생활관심", 1)
ws_life.append(["지역", "업종/카테고리", "구체적장소", "관심사"])
ws_life.append(["배방", "카페", "메가커피", "맛집"])
ws_life.append(["배방", "음식점", "공수리", "가족외식"])
ws_life.append(["아산", "공공기관", "시청", "민원"])
ws_life.append(["천안", "문화", "영화관", "데이트"])
ws_life.append(["", "병원", "", "건강검진"])

# 1-3. 현장 특화 키워드 DB
ws_project = wb.create_sheet("DB_현장특화", 2)
ws_project.append(["현장명", "브랜드명", "경쟁단지", "주요시설"])
ws_project.append(["배방우방아이유쉘", "우방", "", "공수초등학교"])
ws_project.append(["", "아이유쉘", "", "배방역"])
ws_project.append(["", "", "", ""])

# ============================================
# 2. 조합 시트
# ============================================

ws_combo = wb.create_sheet("조합_실행", 3)

# 헤더
headers = [
    "그룹1\n(지역/현장)",
    "그룹2\n(유형/업종)",
    "그룹3\n(특성/관심사)",
    "",
    "▶ 2단조합\n(1+2)",
    "▶ 3단조합\n(1+2+3)",
    "",
    "분류",
    "비고"
]
ws_combo.append(headers)

# 헤더 스타일
for cell in ws_combo[1]:
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 예시 데이터
examples = [
    ["배방", "아파트", "분양", "", "", "", "부동산-고단가", ""],
    ["배방우방아이유쉘", "", "", "", "", "", "브랜드", ""],
    ["배방", "카페", "메가커피", "", "", "", "생활-에펠탑", "노출용"],
    ["아산", "분양", "정보", "", "", "", "부동산-에펠탑", ""],
]

for row_data in examples:
    ws_combo.append(row_data)

# 조합 수식 추가 (5행부터 1000행까지)
for row in range(5, 1001):
    # E열: 2단 조합 (A+B)
    ws_combo[f"E{row}"] = f'=IF(AND(A{row}<>"", B{row}<>""), A{row}&" "&B{row}, "")'
    # F열: 3단 조합 (A+B+C or E+C)
    ws_combo[f"F{row}"] = f'=IF(AND(E{row}<>"", C{row}<>""), E{row}&" "&C{row}, E{row})'

# 컬럼 너비 조정
ws_combo.column_dimensions['A'].width = 18
ws_combo.column_dimensions['B'].width = 18
ws_combo.column_dimensions['C'].width = 18
ws_combo.column_dimensions['D'].width = 2
ws_combo.column_dimensions['E'].width = 25
ws_combo.column_dimensions['F'].width = 30
ws_combo.column_dimensions['G'].width = 2
ws_combo.column_dimensions['H'].width = 15
ws_combo.column_dimensions['I'].width = 15

# ============================================
# 3. 네이버 업로드용 추출 시트
# ============================================

ws_export = wb.create_sheet("네이버_업로드용", 4)
export_headers = ["광고그룹ID", "키워드", "입찰가", "매칭옵션", "사용여부"]
ws_export.append(export_headers)

# 헤더 스타일
for cell in ws_export[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

# 안내 문구
ws_export["G2"] = "📌 사용법:"
ws_export["G3"] = "1. '조합_실행' 시트에서 완성된 키워드(F열) 복사"
ws_export["G4"] = "2. 이 시트의 B열에 붙여넣기"
ws_export["G5"] = "3. 입찰가는 고단가 150~200원, 에펠탑 70원"
ws_export["G6"] = "4. 매칭옵션: 확대 (고정)"
ws_export["G7"] = "5. CSV 저장 시 UTF-8 (BOM) 인코딩 필수"

# ============================================
# 4. 사용 가이드 시트
# ============================================

ws_guide = wb.create_sheet("📖_사용가이드", 5)
ws_guide.column_dimensions['A'].width = 80

guide_content = [
    ["네이버 키워드 조합기 - 사용 가이드"],
    [""],
    ["■ 전체 워크플로우"],
    ["1단계: DB 시트에 키워드 입력 → 2단계: '조합_실행' 시트에서 자동 조합 → 3단계: '네이버_업로드용'에 복사"],
    [""],
    ["■ DB 시트 작성 방법"],
    [""],
    ["📁 DB_부동산:"],
    ["  - 지역, 건물유형, 거래유형, 타겟, 특성 키워드를 행으로 추가"],
    ["  - 예: 배방 | 아파트 | 분양 | 실거주 | 역세권"],
    [""],
    ["📁 DB_생활관심:"],
    ["  - 에펠탑 효과용 저단가 키워드"],
    ["  - 예: 배방 | 카페 | 메가커피 | 맛집"],
    [""],
    ["📁 DB_현장특화:"],
    ["  - 현장명, 브랜드명, 경쟁단지, 주요시설"],
    ["  - 예: 배방우방아이유쉘 | 우방 | - | 공수초등학교"],
    [""],
    ["■ 조합 실행 방법"],
    [""],
    ["📝 조합_실행 시트:"],
    ["  - A열(그룹1): DB에서 복사하거나 직접 입력"],
    ["  - B열(그룹2): 유형이나 업종 키워드"],
    ["  - C열(그룹3): 추가 특성이나 관심사"],
    ["  - E열(2단조합): 자동 계산됨 (A+B)"],
    ["  - F열(3단조합): 자동 계산됨 (A+B+C)"],
    ["  - H열(분류): 용도 표시 (고단가/에펠탑/브랜드)"],
    [""],
    ["■ 네이버 업로드 준비"],
    [""],
    ["📤 네이버_업로드용 시트:"],
    ["  1. 조합_실행의 F열(완성 키워드) 복사"],
    ["  2. 네이버_업로드용의 B열에 붙여넣기"],
    ["  3. C열(입찰가): 브랜드/고단가 150-200원, 에펠탑 70원"],
    ["  4. D열(매칭옵션): '확대' 입력"],
    ["  5. E열(사용여부): 'Y' 입력"],
    ["  6. A열(광고그룹ID): 비워두기 (네이버에서 자동 할당)"],
    [""],
    ["■ CSV 내보내기"],
    ["  - '네이버_업로드용' 시트 전체 선택 후 CSV로 저장"],
    ["  - 인코딩: UTF-8 (BOM) 필수! (한글 깨짐 방지)"],
    ["  - 파일명 예시: 배방우방_키워드_500개_20241128.csv"],
    [""],
    ["■ 키워드 전략"],
    [""],
    ["🎯 고단가 키워드 (150-200원):"],
    ["  - 현장명 직접 언급: 배방우방아이유쉘, 배방 우방"],
    ["  - 전환율 높은 키워드: [현장명] 분양, [현장명] 모델하우스"],
    [""],
    ["🗼 에펠탑 효과 키워드 (70원):"],
    ["  - 지역 생활 키워드: 배방 카페, 배방 맛집"],
    ["  - 반복 노출 목적, 클릭 유도 X, 브랜드 각인 효과"],
    ["  - 광고 그룹 100개 × 키워드 1000개 = 100,000개 세팅 가능"],
    [""],
    ["■ 주의사항"],
    ["  ⚠️ 중복 키워드 제거 필수"],
    ["  ⚠️ 존재하지 않는 조합 삭제 (예: 배방읍 회룡리 아파트)"],
    ["  ⚠️ 네이버 제한: 광고그룹당 키워드 1000개, 파일당 10000행"],
    [""],
]

for row_idx, content in enumerate(guide_content, start=1):
    ws_guide.cell(row=row_idx, column=1, value=content[0])
    if row_idx == 1:
        cell = ws_guide.cell(row=row_idx, column=1)
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# 기본 시트 제거
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# 저장
output_path = "네이버_키워드_조합기_v2.xlsx"
wb.save(output_path)

print(f"[OK] Created: {output_path}")
print(f"Sheets:")
print(f"  1. DB_Real_Estate")
print(f"  2. DB_Life_Interest")
print(f"  3. DB_Project_Specific")
print(f"  4. Combination_Main")
print(f"  5. Naver_Upload")
print(f"  6. User_Guide")
